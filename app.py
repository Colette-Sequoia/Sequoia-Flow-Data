from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import openpyxl
from datetime import datetime
import os
import tempfile
from werkzeug.utils import secure_filename
import traceback
import gc

app = Flask(__name__)
CORS(app)

# Configuration from environment variables with defaults
UPLOAD_FOLDER = tempfile.gettempdir()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_CONTENT_LENGTH', 100 * 1024 * 1024))  # 100MB default

@app.route('/')
def index():
    """Serve the main application page"""
    return send_file('index.html')

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'message': 'Flows processor API is running'})

@app.route('/process', methods=['POST'])
def process_flows():
    try:
        print("=== Starting process request ===")
        
        # Check if all files are present
        required_files = ['returns', 'currentAUM', 'currentFlows', 'previousFlows', 'gla']
        
        for file_key in required_files:
            if file_key not in request.files:
                print(f"Missing file: {file_key}")
                return jsonify({'error': f'Missing file: {file_key}'}), 400
        
        print("All files present, saving to disk...")
        
        # Save uploaded files
        file_paths = {}
        for file_key in required_files:
            file = request.files[file_key]
            if file.filename == '':
                print(f"Empty filename for: {file_key}")
                return jsonify({'error': f'Empty filename for: {file_key}'}), 400
            
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'{file_key}_{filename}')
            file.save(filepath)
            file_paths[file_key] = filepath
            print(f"Saved {file_key}: {filepath}")
        
        # Process the files
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output_flows.xlsx')
        
        print("Starting data processing...")
        result = process_flows_data(
            file_paths['returns'],
            file_paths['currentAUM'],
            file_paths['currentFlows'],
            file_paths['previousFlows'],
            file_paths['gla'],
            output_path
        )
        
        print(f"Processing result: {result.get('success', False)}")
        
        # Clean up input files
        for filepath in file_paths.values():
            try:
                os.remove(filepath)
                print(f"Cleaned up: {filepath}")
            except Exception as cleanup_error:
                print(f"Cleanup warning: {cleanup_error}")
        
        if result['success']:
            # Force garbage collection to free memory
            gc.collect()
            print("Processing completed successfully")
            
            # Return summary data instead of file
            return jsonify({
                'success': True,
                'summary': result['summary'],
                'download_ready': True
            })
        else:
            error_msg = result.get('error', 'Unknown error')
            print(f"Processing failed: {error_msg}")
            return jsonify({'error': error_msg}), 500
            
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"=== ERROR in /process endpoint ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{error_trace}")
        return jsonify({
            'error': str(e),
            'trace': error_trace,
            'message': 'Check Render logs for detailed error information'
        }), 500

@app.route('/download', methods=['GET'])
def download_file():
    """Download the processed file"""
    try:
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output_flows.xlsx')
        
        if not os.path.exists(output_path):
            return jsonify({'error': 'File not found. Please process files first.'}), 404
        
        response = send_file(
            output_path,
            as_attachment=True,
            download_name='Flows_Final.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Schedule cleanup after download
        @response.call_on_close
        def cleanup():
            try:
                os.remove(output_path)
            except:
                pass
            gc.collect()
        
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def process_flows_data(returns_file, current_aum_file, current_flows_file, previous_flows_file, gla_file, output_file):
    try:
        print("=== Starting process_flows_data ===")
        CALC_LISPS = ['Allan Gray', 'Ninety One', 'Stanlib']
        CALC_PRODUCTS = ['Model']
        
        # 1. Read Performance sheet from template - use read_only first
        print("Reading Performance sheet...")
        template_wb_ro = openpyxl.load_workbook(current_flows_file, read_only=True, data_only=True)
        template_perf_ro = template_wb_ro['Performance']
        
        calc_fund_names = []
        performance_dict = {}
        for row in template_perf_ro.iter_rows(min_row=2, max_row=100):
            if row[0].value and row[1].value is not None:
                fund_name = row[0].value
                perf_value = row[1].value
                calc_fund_names.append(fund_name)
                performance_dict[fund_name] = perf_value
        
        template_wb_ro.close()
        del template_wb_ro
        gc.collect()
        print(f"Loaded {len(calc_fund_names)} performance values")
        
        # 2. Read AUM Mappings with chunking for memory efficiency
        print("Reading AUM mappings...")
        advisor_df = pd.read_excel(current_aum_file, sheet_name='ADVISOR ID | CODE MAP', header=0)
        advisor_df.columns = ['Broker_Code', 'Broker_Name', 'Broker_House', 'LISP', 'Data_Source']
        advisor_df = advisor_df[advisor_df['Broker_Code'] != 'Broker ID/Code/Map']
        advisor_df['Broker_Code_Numeric'] = pd.to_numeric(advisor_df['Broker_Code'], errors='coerce')
        
        broker_lookup = {}
        for _, row in advisor_df.iterrows():
            if pd.notna(row['Broker_Code_Numeric']):
                code = int(row['Broker_Code_Numeric'])
                broker_lookup[code] = {
                    'Broker_Name': row['Broker_Name'],
                    'Broker_House': row['Broker_House']
                }
        
        del advisor_df
        gc.collect()
        print(f"Loaded {len(broker_lookup)} broker mappings")
        
        fund_df = pd.read_excel(current_aum_file, sheet_name='FUND MAP')
        fund_mapping = {}
        for _, row in fund_df.iterrows():
            source_name = row.iloc[0]
            target_name = row.iloc[-1]
            if pd.notna(source_name) and pd.notna(target_name):
                fund_mapping[source_name] = target_name
        
        del fund_df
        gc.collect()
        print(f"Loaded {len(fund_mapping)} fund mappings")
        
        # 3. Read Current AUM
        print("Reading Current AUM...")
        aum_df = pd.read_excel(current_aum_file, sheet_name='FINAL')
        aum_df_clean = aum_df[aum_df['Product'] != 'Life Pool'].copy()
        del aum_df
        gc.collect()
        print(f"Loaded {len(aum_df_clean)} AUM records")
        
        # 4. Process GLA with chunking
        print("Processing GLA data...")
        gla_df = pd.read_excel(gla_file, sheet_name='AUA and Flows')
        print(f"Read {len(gla_df)} GLA records")
        
        gla_agg = gla_df.groupby(['BROKER CODE', 'BROKER NAME', 'HOUSE CODE', 'HOUSE NAME', 
                                   'FCODE', 'F NAME'], as_index=False).agg({
            'INFLOWS': 'sum',
            'OUTFLOWS': 'sum',
            'NETFLOWS': 'sum',
            'AUA': 'sum'
        })
        
        del gla_df
        gc.collect()
        print(f"Aggregated to {len(gla_agg)} GLA records")
        
        gla_mapped_rows = []
        for _, row in gla_agg.iterrows():
            broker_code = row['BROKER CODE']
            fund_name_gla = row['F NAME']
            
            if broker_code in broker_lookup:
                broker_info = broker_lookup[broker_code]
                broker_name = broker_info['Broker_Name']
                broker_house = broker_info['Broker_House']
            else:
                broker_name = row['BROKER NAME']
                broker_house = row['HOUSE NAME']
            
            fund_name = fund_mapping.get(fund_name_gla, fund_name_gla)
            
            gla_mapped_rows.append({
                'Date': pd.Timestamp('2026-02-28'),
                'Broker House Name': broker_house,
                'Broker Name': broker_name,
                'Retirement Fund Type': None,
                'Participating Employer': None,
                'Product': 'Life Pool',
                'LISP': 'Momentum',
                'Fund Name': fund_name,
                'InFlows (R)': row['INFLOWS'],
                'OutFlows (R)': row['OUTFLOWS'],
                'NetFlows (R)': row['NETFLOWS'],
                'AUM (R)': row['AUA']
            })
        
        gla_converted = pd.DataFrame(gla_mapped_rows)
        del gla_agg, gla_mapped_rows
        gc.collect()
        print(f"Mapped {len(gla_converted)} GLA records")
        
        # 5. Combine
        print("Combining data...")
        combined_aum = pd.concat([aum_df_clean, gla_converted], ignore_index=True)
        del aum_df_clean, gla_converted
        gc.collect()
        print(f"Total combined records: {len(combined_aum)}")
        
        def needs_calculation(row):
            lisp = row['LISP']
            product = row['Product']
            fund_name = row['Fund Name']
            
            if lisp == 'Investec' and product == 'Note':
                return True
            
            if lisp in CALC_LISPS and product in CALC_PRODUCTS:
                if fund_name in calc_fund_names:
                    return True
            
            return False
        
        combined_aum['NEEDS_CALC'] = combined_aum.apply(needs_calculation, axis=1)
        calculated_count = combined_aum['NEEDS_CALC'].sum()
        print(f"Records needing calculation: {calculated_count}")
        
        # 6. Read Previous Month
        print("Reading previous month data...")
        prev_flows_df = pd.read_excel(previous_flows_file, sheet_name='Worksheet')
        print(f"Loaded {len(prev_flows_df)} previous records")
        
        # 7. Load template for writing
        print("Loading template workbook...")
        flows_wb = openpyxl.load_workbook(current_flows_file, write_only=False)
        
        # 8. Keep Performance sheet
        perf_sheet = flows_wb['Performance']
        
        # 9. Update Prev M Aum
        print("Updating Prev M Aum sheet...")
        prev_aum_sheet = flows_wb['Prev M Aum']
        
        # Clear existing data efficiently
        prev_aum_sheet.delete_rows(2, prev_aum_sheet.max_row)
        
        # Write previous month data
        for idx, row in prev_flows_df.iterrows():
            row_num = idx + 2
            helper_key = f"{row['Broker House Name']}{row['Broker Name']}{row['LISP']}{row['Fund Name']}"
            prev_aum_sheet.append([
                helper_key,
                row['Broker House Name'],
                row['Broker Name'],
                row['Product'],
                None,
                row['LISP'],
                row['Fund Name'],
                row['AUM (R)']
            ])
        
        del prev_flows_df
        gc.collect()
        print("Prev M Aum updated")
        
        # 10. Create Worksheet
        print("Creating Worksheet...")
        worksheet_sheet = flows_wb['Worksheet']
        worksheet_sheet.delete_rows(2, worksheet_sheet.max_row)
        
        for idx, row in combined_aum.iterrows():
            row_num = idx + 2
            
            if row['NEEDS_CALC']:
                worksheet_sheet.append([
                    row['Date'],
                    row['Broker House Name'],
                    row['Broker Name'],
                    row.get('Retirement Fund Type'),
                    row.get('Participating Employer'),
                    row['Product'],
                    row['LISP'],
                    row['Fund Name'],
                    0, 0, 0,
                    row['AUM (R)'],
                    None,
                    f'=B{row_num}&C{row_num}&G{row_num}&H{row_num}',
                    f"=IFERROR(VLOOKUP(N{row_num},'Prev M Aum'!$A$1:$H$1000,8,FALSE),0)",
                    f'=IFERROR(VLOOKUP(H{row_num},Performance!$A$2:$B$100,2,FALSE),0)',
                    f'=O{row_num}',
                    f'=(Q{row_num}*(1+P{row_num}))',
                    f'=L{row_num}-R{row_num}'
                ])
            else:
                worksheet_sheet.append([
                    row['Date'],
                    row['Broker House Name'],
                    row['Broker Name'],
                    row.get('Retirement Fund Type'),
                    row.get('Participating Employer'),
                    row['Product'],
                    row['LISP'],
                    row['Fund Name'],
                    row.get('InFlows (R)', 0),
                    row.get('OutFlows (R)', 0),
                    row.get('NetFlows (R)', 0),
                    row['AUM (R)'],
                    None, None, None, None, None, None,
                    row.get('NetFlows (R)', 0)
                ])
        
        print("Worksheet created")
        
        # 11. Create Final
        print("Creating Final sheet...")
        final_sheet = flows_wb['Final']
        final_sheet.delete_rows(2, final_sheet.max_row)
        
        worksheet_sheet = flows_wb['Worksheet']
        for row_idx, row in enumerate(worksheet_sheet.iter_rows(min_row=2, max_row=len(combined_aum)+1), start=2):
            needs_calc_val = combined_aum.iloc[row_idx-2]['NEEDS_CALC']
            
            final_sheet.append([
                row[0].value,  # Date
                row[1].value,  # Broker House Name
                row[2].value,  # Broker Name
                row[3].value,  # Retirement Fund Type
                row[4].value,  # Participating Employer
                row[5].value,  # Product
                row[6].value,  # LISP
                row[7].value,  # Fund Name
                row[8].value,  # InFlows
                row[9].value,  # OutFlows
                f'=Worksheet!S{row_idx}' if needs_calc_val else row[10].value  # NetFlows
            ])
        
        print("Final sheet created")
        
        # 12. Save
        print("Saving workbook...")
        flows_wb.save(output_file)
        flows_wb.close()
        del flows_wb
        gc.collect()
        print("Workbook saved")
        
        # Calculate summary statistics - read back to avoid keeping workbook in memory
        print("Calculating summary...")
        final_df = pd.read_excel(output_file, sheet_name='Final')
        
        summary = {
            'total_netflows': float(final_df['NetFlows (R)'].sum()),
            'total_rows': len(final_df),
            'by_lisp': {},
            'calculated_rows': int(calculated_count),
            'copied_rows': len(combined_aum) - int(calculated_count)
        }
        
        # Calculate by LISP
        lisp_summary = final_df.groupby('LISP').agg({
            'NetFlows (R)': 'sum',
            'LISP': 'count'
        }).rename(columns={'LISP': 'count'})
        
        for lisp, row in lisp_summary.iterrows():
            summary['by_lisp'][lisp] = {
                'netflows': float(row['NetFlows (R)']),
                'count': int(row['count'])
            }
        
        del final_df, lisp_summary
        gc.collect()
        
        # Calculate formulas manually
        print("Resolving formulas...")
        wb = openpyxl.load_workbook(output_file)
        worksheet = wb['Worksheet']
        perf_sheet = wb['Performance']
        prev_aum_sheet = wb['Prev M Aum']
        
        perf_lookup = {r[0].value: r[1].value for r in perf_sheet.iter_rows(min_row=2, max_row=100) if r[0].value}
        prev_aum_lookup = {r[0].value: r[7].value for r in prev_aum_sheet.iter_rows(min_row=2, max_row=1000) if r[0].value}
        
        for row_num in range(2, worksheet.max_row + 1):
            s_cell = worksheet[f'S{row_num}']
            if isinstance(s_cell.value, str) and s_cell.value.startswith('='):
                b = worksheet[f'B{row_num}'].value
                c = worksheet[f'C{row_num}'].value
                g = worksheet[f'G{row_num}'].value
                h = worksheet[f'H{row_num}'].value
                l = worksheet[f'L{row_num}'].value
                
                if b and c and g and h:
                    key = f"{b}{c}{g}{h}"
                    prev = prev_aum_lookup.get(key, 0)
                    perf = perf_lookup.get(h, 0)
                    adj = prev * (1 + perf)
                    nf = l - adj if l else 0
                    
                    worksheet[f'S{row_num}'].value = nf
                    worksheet[f'O{row_num}'].value = prev
                    worksheet[f'P{row_num}'].value = perf
                    worksheet[f'Q{row_num}'].value = prev
                    worksheet[f'R{row_num}'].value = adj
        
        final_sheet = wb['Final']
        for row_num in range(2, final_sheet.max_row + 1):
            k = final_sheet[f'K{row_num}']
            if isinstance(k.value, str) and '=Worksheet!' in k.value:
                final_sheet[f'K{row_num}'].value = worksheet[f'S{row_num}'].value
        
        wb.save(output_file)
        wb.close()
        del wb, combined_aum
        gc.collect()
        
        print("=== Processing complete ===")
        return {'success': True, 'summary': summary}
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"=== ERROR in process_flows_data ===")
        print(f"Error: {str(e)}")
        print(error_trace)
        return {'success': False, 'error': f"{str(e)}\n\nTraceback:\n{error_trace}"}
        
        # 1. Read Performance sheet from template
        template_wb = openpyxl.load_workbook(current_flows_file, read_only=False, data_only=False)
        template_perf = template_wb['Performance']
        
        calc_fund_names = []
        performance_dict = {}
        for row in template_perf.iter_rows(min_row=2, max_row=100):
            if row[0].value and row[1].value is not None:
                fund_name = row[0].value
                perf_value = row[1].value
                calc_fund_names.append(fund_name)
                performance_dict[fund_name] = perf_value
        
        # 2. Read AUM Mappings with chunking for memory efficiency
        advisor_df = pd.read_excel(current_aum_file, sheet_name='ADVISOR ID | CODE MAP', header=0)
        advisor_df.columns = ['Broker_Code', 'Broker_Name', 'Broker_House', 'LISP', 'Data_Source']
        advisor_df = advisor_df[advisor_df['Broker_Code'] != 'Broker ID/Code/Map']
        advisor_df['Broker_Code_Numeric'] = pd.to_numeric(advisor_df['Broker_Code'], errors='coerce')
        
        broker_lookup = {}
        for _, row in advisor_df.iterrows():
            if pd.notna(row['Broker_Code_Numeric']):
                code = int(row['Broker_Code_Numeric'])
                broker_lookup[code] = {
                    'Broker_Name': row['Broker_Name'],
                    'Broker_House': row['Broker_House']
                }
        
        # Free memory
        del advisor_df
        gc.collect()
        
        fund_df = pd.read_excel(current_aum_file, sheet_name='FUND MAP')
        fund_mapping = {}
        for _, row in fund_df.iterrows():
            source_name = row.iloc[0]
            target_name = row.iloc[-1]
            if pd.notna(source_name) and pd.notna(target_name):
                fund_mapping[source_name] = target_name
        
        del fund_df
        gc.collect()
        
        # 3. Read Current AUM
        aum_df = pd.read_excel(current_aum_file, sheet_name='FINAL')
        aum_df_clean = aum_df[aum_df['Product'] != 'Life Pool'].copy()
        del aum_df
        gc.collect()
        
        # 4. Process GLA with chunking
        gla_df = pd.read_excel(gla_file, sheet_name='AUA and Flows')
        gla_agg = gla_df.groupby(['BROKER CODE', 'BROKER NAME', 'HOUSE CODE', 'HOUSE NAME', 
                                   'FCODE', 'F NAME']).agg({
            'INFLOWS': 'sum',
            'OUTFLOWS': 'sum',
            'NETFLOWS': 'sum',
            'AUA': 'sum'
        }).reset_index()
        
        del gla_df
        gc.collect()
        
        gla_mapped_rows = []
        for _, row in gla_agg.iterrows():
            broker_code = row['BROKER CODE']
            fund_name_gla = row['F NAME']
            
            if broker_code in broker_lookup:
                broker_info = broker_lookup[broker_code]
                broker_name = broker_info['Broker_Name']
                broker_house = broker_info['Broker_House']
            else:
                broker_name = row['BROKER NAME']
                broker_house = row['HOUSE NAME']
            
            fund_name = fund_mapping.get(fund_name_gla, fund_name_gla)
            
            gla_mapped_rows.append({
                'Date': pd.Timestamp('2026-02-28'),
                'Broker House Name': broker_house,
                'Broker Name': broker_name,
                'Retirement Fund Type': None,
                'Participating Employer': None,
                'Product': 'Life Pool',
                'LISP': 'Momentum',
                'Fund Name': fund_name,
                'InFlows (R)': row['INFLOWS'],
                'OutFlows (R)': row['OUTFLOWS'],
                'NetFlows (R)': row['NETFLOWS'],
                'AUM (R)': row['AUA']
            })
        
        gla_converted = pd.DataFrame(gla_mapped_rows)
        del gla_agg, gla_mapped_rows
        gc.collect()
        
        # 5. Combine
        combined_aum = pd.concat([aum_df_clean, gla_converted], ignore_index=True)
        del aum_df_clean, gla_converted
        gc.collect()
        
        def needs_calculation(row):
            lisp = row['LISP']
            product = row['Product']
            fund_name = row['Fund Name']
            
            if lisp == 'Investec' and product == 'Note':
                return True
            
            if lisp in CALC_LISPS and product in CALC_PRODUCTS:
                if fund_name in calc_fund_names:
                    return True
            
            return False
        
        combined_aum['NEEDS_CALC'] = combined_aum.apply(needs_calculation, axis=1)
        
        # 6. Read Previous Month
        prev_flows_df = pd.read_excel(previous_flows_file, sheet_name='Worksheet')
        
        # 7. Load template
        flows_wb = openpyxl.load_workbook(current_flows_file)
        
        # 8. Keep Performance sheet
        perf_sheet = flows_wb['Performance']
        
        # 9. Update Prev M Aum
        prev_aum_sheet = flows_wb['Prev M Aum']
        for row in prev_aum_sheet.iter_rows(min_row=2, max_row=prev_aum_sheet.max_row):
            for cell in row:
                cell.value = None
        
        row_num = 2
        for _, row in prev_flows_df.iterrows():
            helper_key = f"{row['Broker House Name']}{row['Broker Name']}{row['LISP']}{row['Fund Name']}"
            prev_aum_sheet[f'A{row_num}'] = helper_key
            prev_aum_sheet[f'B{row_num}'] = row['Broker House Name']
            prev_aum_sheet[f'C{row_num}'] = row['Broker Name']
            prev_aum_sheet[f'D{row_num}'] = row['Product']
            prev_aum_sheet[f'F{row_num}'] = row['LISP']
            prev_aum_sheet[f'G{row_num}'] = row['Fund Name']
            prev_aum_sheet[f'H{row_num}'] = row['AUM (R)']
            row_num += 1
        
        del prev_flows_df
        gc.collect()
        
        # 10. Create Worksheet
        worksheet_sheet = flows_wb['Worksheet']
        for row in worksheet_sheet.iter_rows(min_row=2, max_row=worksheet_sheet.max_row):
            for cell in row:
                cell.value = None
        
        row_num = 2
        for _, row in combined_aum.iterrows():
            worksheet_sheet[f'A{row_num}'] = row['Date']
            worksheet_sheet[f'B{row_num}'] = row['Broker House Name']
            worksheet_sheet[f'C{row_num}'] = row['Broker Name']
            worksheet_sheet[f'D{row_num}'] = row.get('Retirement Fund Type')
            worksheet_sheet[f'E{row_num}'] = row.get('Participating Employer')
            worksheet_sheet[f'F{row_num}'] = row['Product']
            worksheet_sheet[f'G{row_num}'] = row['LISP']
            worksheet_sheet[f'H{row_num}'] = row['Fund Name']
            
            if row['NEEDS_CALC']:
                worksheet_sheet[f'I{row_num}'] = 0
                worksheet_sheet[f'J{row_num}'] = 0
                worksheet_sheet[f'K{row_num}'] = 0
                worksheet_sheet[f'L{row_num}'] = row['AUM (R)']
                worksheet_sheet[f'N{row_num}'] = f'=B{row_num}&C{row_num}&G{row_num}&H{row_num}'
                worksheet_sheet[f'O{row_num}'] = f"=IFERROR(VLOOKUP(N{row_num},'Prev M Aum'!$A$1:$H$1000,8,FALSE),0)"
                worksheet_sheet[f'P{row_num}'] = f'=IFERROR(VLOOKUP(H{row_num},Performance!$A$2:$B$100,2,FALSE),0)'
                worksheet_sheet[f'Q{row_num}'] = f'=O{row_num}'
                worksheet_sheet[f'R{row_num}'] = f'=(Q{row_num}*(1+P{row_num}))'
                worksheet_sheet[f'S{row_num}'] = f'=L{row_num}-R{row_num}'
            else:
                worksheet_sheet[f'I{row_num}'] = row.get('InFlows (R)', 0)
                worksheet_sheet[f'J{row_num}'] = row.get('OutFlows (R)', 0)
                worksheet_sheet[f'K{row_num}'] = row.get('NetFlows (R)', 0)
                worksheet_sheet[f'L{row_num}'] = row['AUM (R)']
                worksheet_sheet[f'S{row_num}'] = row.get('NetFlows (R)', 0)
            
            row_num += 1
        
        # 11. Create Final
        final_sheet = flows_wb['Final']
        for row in final_sheet.iter_rows(min_row=2, max_row=final_sheet.max_row):
            for cell in row:
                cell.value = None
        
        row_num = 2
        for ws_row in worksheet_sheet.iter_rows(min_row=2, max_row=worksheet_sheet.max_row):
            if ws_row[0].value is None:
                break
            
            for col_idx in range(10):
                final_sheet.cell(row=row_num, column=col_idx+1).value = ws_row[col_idx].value
            
            needs_calc_val = combined_aum.iloc[row_num-2]['NEEDS_CALC']
            if needs_calc_val:
                final_sheet.cell(row=row_num, column=11).value = f'=Worksheet!S{row_num}'
            else:
                final_sheet.cell(row=row_num, column=11).value = ws_row[10].value
            
            row_num += 1
        
        # 12. Save
        flows_wb.save(output_file)
        del flows_wb, template_wb
        gc.collect()
        
        # Calculate summary statistics - read back to avoid keeping workbook in memory
        final_df = pd.read_excel(output_file, sheet_name='Final')
        
        summary = {
            'total_netflows': float(final_df['NetFlows (R)'].sum()),
            'total_rows': len(final_df),
            'by_lisp': {},
            'calculated_rows': len(combined_aum[combined_aum['NEEDS_CALC'] == True]),
            'copied_rows': len(combined_aum[combined_aum['NEEDS_CALC'] == False])
        }
        
        # Calculate by LISP
        lisp_summary = final_df.groupby('LISP').agg({
            'NetFlows (R)': 'sum',
            'LISP': 'count'
        }).rename(columns={'LISP': 'count'})
        
        for lisp, row in lisp_summary.iterrows():
            summary['by_lisp'][lisp] = {
                'netflows': float(row['NetFlows (R)']),
                'count': int(row['count'])
            }
        
        # Calculate formulas manually
        wb = openpyxl.load_workbook(output_file)
        worksheet = wb['Worksheet']
        perf_sheet = wb['Performance']
        prev_aum_sheet = wb['Prev M Aum']
        
        perf_lookup = {r[0].value: r[1].value for r in perf_sheet.iter_rows(min_row=2, max_row=100) if r[0].value}
        prev_aum_lookup = {r[0].value: r[7].value for r in prev_aum_sheet.iter_rows(min_row=2, max_row=1000) if r[0].value}
        
        for row_num in range(2, worksheet.max_row + 1):
            s_cell = worksheet[f'S{row_num}']
            if isinstance(s_cell.value, str) and s_cell.value.startswith('='):
                b = worksheet[f'B{row_num}'].value
                c = worksheet[f'C{row_num}'].value
                g = worksheet[f'G{row_num}'].value
                h = worksheet[f'H{row_num}'].value
                l = worksheet[f'L{row_num}'].value
                
                if b and c and g and h:
                    key = f"{b}{c}{g}{h}"
                    prev = prev_aum_lookup.get(key, 0)
                    perf = perf_lookup.get(h, 0)
                    adj = prev * (1 + perf)
                    nf = l - adj if l else 0
                    
                    worksheet[f'S{row_num}'].value = nf
                    worksheet[f'O{row_num}'].value = prev
                    worksheet[f'P{row_num}'].value = perf
                    worksheet[f'Q{row_num}'].value = prev
                    worksheet[f'R{row_num}'].value = adj
        
        final_sheet = wb['Final']
        for row_num in range(2, final_sheet.max_row + 1):
            k = final_sheet[f'K{row_num}']
            if isinstance(k.value, str) and '=Worksheet!' in k.value:
                final_sheet[f'K{row_num}'].value = worksheet[f'S{row_num}'].value
        
        wb.save(output_file)
        del wb, final_df, combined_aum
        gc.collect()
        
        return {'success': True, 'summary': summary}
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in process_flows_data: {str(e)}")
        print(error_trace)
        return {'success': False, 'error': f"{str(e)}\n\nTraceback:\n{error_trace}"}
        CALC_LISPS = ['Allan Gray', 'Ninety One', 'Stanlib']
        CALC_PRODUCTS = ['Model']
        
        # 1. Read Performance sheet from template
        template_wb = openpyxl.load_workbook(current_flows_file)
        template_perf = template_wb['Performance']
        
        calc_fund_names = []
        performance_dict = {}
        for row in template_perf.iter_rows(min_row=2, max_row=100):
            if row[0].value and row[1].value is not None:
                fund_name = row[0].value
                perf_value = row[1].value
                calc_fund_names.append(fund_name)
                performance_dict[fund_name] = perf_value
        
        # 2. Read AUM Mappings
        advisor_df = pd.read_excel(current_aum_file, sheet_name='ADVISOR ID | CODE MAP', header=0)
        advisor_df.columns = ['Broker_Code', 'Broker_Name', 'Broker_House', 'LISP', 'Data_Source']
        advisor_df = advisor_df[advisor_df['Broker_Code'] != 'Broker ID/Code/Map']
        advisor_df['Broker_Code_Numeric'] = pd.to_numeric(advisor_df['Broker_Code'], errors='coerce')
        
        broker_lookup = {}
        for _, row in advisor_df.iterrows():
            if pd.notna(row['Broker_Code_Numeric']):
                code = int(row['Broker_Code_Numeric'])
                broker_lookup[code] = {
                    'Broker_Name': row['Broker_Name'],
                    'Broker_House': row['Broker_House']
                }
        
        fund_df = pd.read_excel(current_aum_file, sheet_name='FUND MAP')
        fund_mapping = {}
        for _, row in fund_df.iterrows():
            source_name = row.iloc[0]
            target_name = row.iloc[-1]
            if pd.notna(source_name) and pd.notna(target_name):
                fund_mapping[source_name] = target_name
        
        # 3. Read Current AUM
        aum_df = pd.read_excel(current_aum_file, sheet_name='FINAL')
        aum_df_clean = aum_df[aum_df['Product'] != 'Life Pool'].copy()
        
        # 4. Process GLA
        gla_df = pd.read_excel(gla_file, sheet_name='AUA and Flows')
        gla_agg = gla_df.groupby(['BROKER CODE', 'BROKER NAME', 'HOUSE CODE', 'HOUSE NAME', 
                                   'FCODE', 'F NAME']).agg({
            'INFLOWS': 'sum',
            'OUTFLOWS': 'sum',
            'NETFLOWS': 'sum',
            'AUA': 'sum'
        }).reset_index()
        
        gla_mapped_rows = []
        for _, row in gla_agg.iterrows():
            broker_code = row['BROKER CODE']
            fund_name_gla = row['F NAME']
            
            if broker_code in broker_lookup:
                broker_info = broker_lookup[broker_code]
                broker_name = broker_info['Broker_Name']
                broker_house = broker_info['Broker_House']
            else:
                broker_name = row['BROKER NAME']
                broker_house = row['HOUSE NAME']
            
            fund_name = fund_mapping.get(fund_name_gla, fund_name_gla)
            
            gla_mapped_rows.append({
                'Date': pd.Timestamp('2026-02-28'),
                'Broker House Name': broker_house,
                'Broker Name': broker_name,
                'Retirement Fund Type': None,
                'Participating Employer': None,
                'Product': 'Life Pool',
                'LISP': 'Momentum',
                'Fund Name': fund_name,
                'InFlows (R)': row['INFLOWS'],
                'OutFlows (R)': row['OUTFLOWS'],
                'NetFlows (R)': row['NETFLOWS'],
                'AUM (R)': row['AUA']
            })
        
        gla_converted = pd.DataFrame(gla_mapped_rows)
        
        # 5. Combine
        combined_aum = pd.concat([aum_df_clean, gla_converted], ignore_index=True)
        
        def needs_calculation(row):
            lisp = row['LISP']
            product = row['Product']
            fund_name = row['Fund Name']
            
            if lisp == 'Investec' and product == 'Note':
                return True
            
            if lisp in CALC_LISPS and product in CALC_PRODUCTS:
                if fund_name in calc_fund_names:
                    return True
            
            return False
        
        combined_aum['NEEDS_CALC'] = combined_aum.apply(needs_calculation, axis=1)
        
        # 6. Read Previous Month
        prev_flows_df = pd.read_excel(previous_flows_file, sheet_name='Worksheet')
        
        # 7. Load template
        flows_wb = openpyxl.load_workbook(current_flows_file)
        
        # 8. Keep Performance sheet
        perf_sheet = flows_wb['Performance']
        
        # 9. Update Prev M Aum
        prev_aum_sheet = flows_wb['Prev M Aum']
        for row in prev_aum_sheet.iter_rows(min_row=2, max_row=prev_aum_sheet.max_row):
            for cell in row:
                cell.value = None
        
        row_num = 2
        for _, row in prev_flows_df.iterrows():
            helper_key = f"{row['Broker House Name']}{row['Broker Name']}{row['LISP']}{row['Fund Name']}"
            prev_aum_sheet[f'A{row_num}'] = helper_key
            prev_aum_sheet[f'B{row_num}'] = row['Broker House Name']
            prev_aum_sheet[f'C{row_num}'] = row['Broker Name']
            prev_aum_sheet[f'D{row_num}'] = row['Product']
            prev_aum_sheet[f'F{row_num}'] = row['LISP']
            prev_aum_sheet[f'G{row_num}'] = row['Fund Name']
            prev_aum_sheet[f'H{row_num}'] = row['AUM (R)']
            row_num += 1
        
        # 10. Create Worksheet
        worksheet_sheet = flows_wb['Worksheet']
        for row in worksheet_sheet.iter_rows(min_row=2, max_row=worksheet_sheet.max_row):
            for cell in row:
                cell.value = None
        
        row_num = 2
        for _, row in combined_aum.iterrows():
            worksheet_sheet[f'A{row_num}'] = row['Date']
            worksheet_sheet[f'B{row_num}'] = row['Broker House Name']
            worksheet_sheet[f'C{row_num}'] = row['Broker Name']
            worksheet_sheet[f'D{row_num}'] = row.get('Retirement Fund Type')
            worksheet_sheet[f'E{row_num}'] = row.get('Participating Employer')
            worksheet_sheet[f'F{row_num}'] = row['Product']
            worksheet_sheet[f'G{row_num}'] = row['LISP']
            worksheet_sheet[f'H{row_num}'] = row['Fund Name']
            
            if row['NEEDS_CALC']:
                worksheet_sheet[f'I{row_num}'] = 0
                worksheet_sheet[f'J{row_num}'] = 0
                worksheet_sheet[f'K{row_num}'] = 0
                worksheet_sheet[f'L{row_num}'] = row['AUM (R)']
                worksheet_sheet[f'N{row_num}'] = f'=B{row_num}&C{row_num}&G{row_num}&H{row_num}'
                worksheet_sheet[f'O{row_num}'] = f"=IFERROR(VLOOKUP(N{row_num},'Prev M Aum'!$A$1:$H$1000,8,FALSE),0)"
                worksheet_sheet[f'P{row_num}'] = f'=IFERROR(VLOOKUP(H{row_num},Performance!$A$2:$B$100,2,FALSE),0)'
                worksheet_sheet[f'Q{row_num}'] = f'=O{row_num}'
                worksheet_sheet[f'R{row_num}'] = f'=(Q{row_num}*(1+P{row_num}))'
                worksheet_sheet[f'S{row_num}'] = f'=L{row_num}-R{row_num}'
            else:
                worksheet_sheet[f'I{row_num}'] = row.get('InFlows (R)', 0)
                worksheet_sheet[f'J{row_num}'] = row.get('OutFlows (R)', 0)
                worksheet_sheet[f'K{row_num}'] = row.get('NetFlows (R)', 0)
                worksheet_sheet[f'L{row_num}'] = row['AUM (R)']
                worksheet_sheet[f'S{row_num}'] = row.get('NetFlows (R)', 0)
            
            row_num += 1
        
        # 11. Create Final
        final_sheet = flows_wb['Final']
        for row in final_sheet.iter_rows(min_row=2, max_row=final_sheet.max_row):
            for cell in row:
                cell.value = None
        
        row_num = 2
        for ws_row in worksheet_sheet.iter_rows(min_row=2, max_row=worksheet_sheet.max_row):
            if ws_row[0].value is None:
                break
            
            for col_idx in range(10):
                final_sheet.cell(row=row_num, column=col_idx+1).value = ws_row[col_idx].value
            
            needs_calc_val = combined_aum.iloc[row_num-2]['NEEDS_CALC']
            if needs_calc_val:
                final_sheet.cell(row=row_num, column=11).value = f'=Worksheet!S{row_num}'
            else:
                final_sheet.cell(row=row_num, column=11).value = ws_row[10].value
            
            row_num += 1
        
        # 12. Save
        flows_wb.save(output_file)
        
        # Calculate formulas manually
        wb = openpyxl.load_workbook(output_file)
        worksheet = wb['Worksheet']
        perf_sheet = wb['Performance']
        prev_aum_sheet = wb['Prev M Aum']
        
        perf_lookup = {r[0].value: r[1].value for r in perf_sheet.iter_rows(min_row=2, max_row=100) if r[0].value}
        prev_aum_lookup = {r[0].value: r[7].value for r in prev_aum_sheet.iter_rows(min_row=2, max_row=1000) if r[0].value}
        
        for row_num in range(2, worksheet.max_row + 1):
            s_cell = worksheet[f'S{row_num}']
            if isinstance(s_cell.value, str) and s_cell.value.startswith('='):
                b = worksheet[f'B{row_num}'].value
                c = worksheet[f'C{row_num}'].value
                g = worksheet[f'G{row_num}'].value
                h = worksheet[f'H{row_num}'].value
                l = worksheet[f'L{row_num}'].value
                
                if b and c and g and h:
                    key = f"{b}{c}{g}{h}"
                    prev = prev_aum_lookup.get(key, 0)
                    perf = perf_lookup.get(h, 0)
                    adj = prev * (1 + perf)
                    nf = l - adj if l else 0
                    
                    worksheet[f'S{row_num}'].value = nf
                    worksheet[f'O{row_num}'].value = prev
                    worksheet[f'P{row_num}'].value = perf
                    worksheet[f'Q{row_num}'].value = prev
                    worksheet[f'R{row_num}'].value = adj
        
        final_sheet = wb['Final']
        for row_num in range(2, final_sheet.max_row + 1):
            k = final_sheet[f'K{row_num}']
            if isinstance(k.value, str) and '=Worksheet!' in k.value:
                final_sheet[f'K{row_num}'].value = worksheet[f'S{row_num}'].value
        
        wb.save(output_file)
        
        # Calculate summary statistics
        final_df = pd.read_excel(output_file, sheet_name='Final')
        
        summary = {
            'total_netflows': float(final_df['NetFlows (R)'].sum()),
            'total_rows': len(final_df),
            'by_lisp': {},
            'calculated_rows': len(combined_aum[combined_aum['NEEDS_CALC'] == True]),
            'copied_rows': len(combined_aum[combined_aum['NEEDS_CALC'] == False])
        }
        
        # Calculate by LISP
        lisp_summary = final_df.groupby('LISP').agg({
            'NetFlows (R)': 'sum',
            'LISP': 'count'
        }).rename(columns={'LISP': 'count'})
        
        for lisp, row in lisp_summary.iterrows():
            summary['by_lisp'][lisp] = {
                'netflows': float(row['NetFlows (R)']),
                'count': int(row['count'])
            }
        
        return {'success': True, 'summary': summary}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
